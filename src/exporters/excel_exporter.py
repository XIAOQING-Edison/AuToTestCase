"""
Excel导出器模块
负责将测试用例导出为格式化的Excel文件，支持样式美化和自动调整
"""

from typing import List
import pandas as pd
from src.core.test_generator import TestCase
from src.config import EXCEL_TEMPLATE_HEADERS, OUTPUT_DIR
import os
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

class ExcelExporter:
    def export(self, test_cases: List[TestCase], output_filename: str) -> str:
        """
        Export test cases to Excel format
        
        Args:
            test_cases (List[TestCase]): List of test cases to export
            output_filename (str): Name of the output file
            
        Returns:
            str: Path to the exported file
        """
        # Convert test cases to DataFrame format
        data = []
        for test_case in test_cases:
            # 处理前置条件
            preconditions_str = ""
            if test_case.preconditions:
                preconditions_str = "\n".join([f"{i+1}. {p}" for i, p in enumerate(test_case.preconditions)])
            
            # 处理测试步骤和预期结果
            steps_str = ""
            expected_results_str = ""
            
            for step in test_case.steps:
                if steps_str:
                    steps_str += "\n"
                steps_str += f"{step.step_number}. {step.description}"
                
                if expected_results_str:
                    expected_results_str += "\n"
                expected_results_str += f"{step.step_number}. {step.expected_result}"
            
            # 添加到数据列表
            data.append({
                "用例编号": test_case.id,
                "所属模块": test_case.module,
                "用例标题": test_case.title,
                "优先级": test_case.priority,
                "前置条件": preconditions_str,
                "测试步骤": steps_str,
                "预期结果": expected_results_str
            })
        
        df = pd.DataFrame(data)
        print(f"\nExporting {len(data)} test cases to Excel")
        print("DataFrame columns:", df.columns.tolist())
        print("DataFrame shape:", df.shape)
        
        # Ensure columns are in the correct order
        df = df.reindex(columns=EXCEL_TEMPLATE_HEADERS)
        
        # Create output path
        if not output_filename.endswith('.xlsx'):
            output_filename += '.xlsx'
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        
        # Export to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Test Cases')
            
            # Get the workbook and the worksheet
            workbook = writer.book
            worksheet = writer.sheets['Test Cases']
            
            # Define styles
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            cell_alignment = Alignment(vertical='top', wrap_text=True)
            
            # Set column widths
            column_widths = {
                'A': 15,  # 用例编号
                'B': 15,  # 所属模块
                'C': 25,  # 用例标题
                'D': 10,  # 优先级
                'E': 20,  # 前置条件
                'F': 40,  # 测试步骤
                'G': 40   # 预期结果
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # Format headers
            for col_num, header in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = header_alignment
            
            # Format all data cells
            for row_num in range(2, len(df) + 2):
                # Set row height
                worksheet.row_dimensions[row_num].height = 60
                
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = border
                    cell.alignment = cell_alignment
                    
                    # Color code priorities
                    if col_num == 4:  # Priority column
                        value = cell.value
                        if value == "高":
                            cell.fill = PatternFill(start_color='FFD9D9', end_color='FFD9D9', fill_type='solid')
                        elif value == "中":
                            cell.fill = PatternFill(start_color='FFFDCC', end_color='FFFDCC', fill_type='solid')
                        elif value == "低":
                            cell.fill = PatternFill(start_color='D9FFD9', end_color='D9FFD9', fill_type='solid')
        
        print(f"Excel file generated: {output_path}")
        return output_path 